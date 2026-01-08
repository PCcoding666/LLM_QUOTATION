"""
Excel导出服务测试
"""
import pytest
from decimal import Decimal
from datetime import datetime, timedelta
from uuid import uuid4
from io import BytesIO
from openpyxl import load_workbook

from app.services.excel_exporter import ExcelExporter, get_excel_exporter
from app.models.quote import QuoteSheet, QuoteItem


class MockQuoteSheet:
    """模拟报价单对象"""
    def __init__(self, **kwargs):
        self.quote_id = kwargs.get('quote_id', uuid4())
        self.quote_no = kwargs.get('quote_no', 'QT202601080001')
        self.customer_name = kwargs.get('customer_name', '测试客户')
        self.project_name = kwargs.get('project_name', '测试项目')
        self.status = kwargs.get('status', 'draft')
        self.total_amount = kwargs.get('total_amount', Decimal('10000.00'))
        self.currency = kwargs.get('currency', 'CNY')
        self.created_at = kwargs.get('created_at', datetime.now())
        self.valid_until = kwargs.get('valid_until', datetime.now() + timedelta(days=30))


class MockQuoteItem:
    """模拟报价项对象"""
    def __init__(self, **kwargs):
        self.item_id = kwargs.get('item_id', uuid4())
        self.product_name = kwargs.get('product_name', 'qwen-max')
        self.spec_config = kwargs.get('spec_config', {'model': 'qwen-max', 'context': '128K'})
        self.quantity = kwargs.get('quantity', 1)
        self.duration_months = kwargs.get('duration_months', 1)
        self.unit_price = kwargs.get('unit_price', Decimal('0.04'))
        self.subtotal = kwargs.get('subtotal', Decimal('1000.00'))
        self.discount_info = kwargs.get('discount_info', None)


class TestExcelExporter:
    """Excel导出服务测试"""
    
    def setup_method(self):
        """初始化测试"""
        self.exporter = ExcelExporter()
    
    @pytest.mark.asyncio
    async def test_generate_standard_quote(self):
        """测试生成标准报价单"""
        # 准备测试数据
        quote = MockQuoteSheet()
        items = [
            MockQuoteItem(
                product_name='qwen-max',
                quantity=1,
                subtotal=Decimal('1000.00')
            ),
            MockQuoteItem(
                product_name='qwen-plus',
                quantity=2,
                subtotal=Decimal('500.00')
            )
        ]
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        
        # 验证生成的文件
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        
        # 读取并验证内容
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 验证标题
        assert ws.title == "报价单"
        assert "阿里云产品报价单" in str(ws['A1'].value)
        
        # 验证客户信息
        customer_found = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value == quote.customer_name:
                    customer_found = True
                    break
        assert customer_found, "未找到客户名称"
    
    @pytest.mark.asyncio
    async def test_generate_simplified_quote(self):
        """测试生成简化版报价单"""
        quote = MockQuoteSheet(total_amount=Decimal('5000.00'))
        items = [
            MockQuoteItem(
                product_name='测试产品',
                quantity=5,
                subtotal=Decimal('5000.00')
            )
        ]
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_simplified_quote(quote, items)
        
        # 验证
        assert excel_bytes is not None
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        assert ws.title == "简化报价单"
        
        # 验证表头
        assert ws['A1'].value == "产品名称"
        assert ws['B1'].value == "数量"
        assert ws['C1'].value == "价格(元)"
        
        # 验证数据行
        assert ws['A2'].value == "测试产品"
        assert ws['B2'].value == 5
    
    @pytest.mark.asyncio
    async def test_empty_items_list(self):
        """测试空商品列表"""
        quote = MockQuoteSheet(total_amount=Decimal('0.00'))
        items = []
        
        # 生成Excel
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
    
    @pytest.mark.asyncio
    async def test_competitor_comparison(self):
        """测试竞品对比版（当前返回标准版）"""
        quote = MockQuoteSheet()
        items = [MockQuoteItem()]
        competitor_data = {"competitor": "test"}
        
        excel_bytes = await self.exporter.generate_competitor_comparison(
            quote, items, competitor_data
        )
        
        assert excel_bytes is not None
    
    def test_format_spec_config(self):
        """测试规格配置格式化"""
        # 测试有效配置
        spec_config = {
            'model': 'qwen-max',
            'context': '128K',
            'region': 'cn-beijing',  # 应该被过滤
            'spec_type': 'standard'  # 应该被过滤
        }
        result = self.exporter._format_spec_config(spec_config)
        assert 'model' in result
        assert 'context' in result
        assert 'region' not in result
        
        # 测试空配置
        assert self.exporter._format_spec_config(None) == "-"
        assert self.exporter._format_spec_config({}) == "-"
    
    def test_format_discount_info(self):
        """测试折扣信息格式化"""
        # 测试有效折扣
        discount_info = {
            'discounts': [
                {'type': 'tiered', 'value': 9},
                {'type': 'batch', 'value': 5}
            ]
        }
        result = self.exporter._format_discount_info(discount_info)
        assert '阶梯折扣' in result
        assert 'Batch折扣' in result
        
        # 测试空折扣
        assert self.exporter._format_discount_info(None) == "-"
        assert self.exporter._format_discount_info({}) == "-"
    
    def test_get_excel_exporter_singleton(self):
        """测试获取导出器单例"""
        exporter1 = get_excel_exporter()
        exporter2 = get_excel_exporter()
        
        assert exporter1 is exporter2


class TestExcelExporterEdgeCases:
    """Excel导出器边界情况测试"""
    
    def setup_method(self):
        self.exporter = ExcelExporter()
    
    @pytest.mark.asyncio
    async def test_large_amount(self):
        """测试大金额数值"""
        quote = MockQuoteSheet(total_amount=Decimal('9999999999.99'))
        items = [
            MockQuoteItem(
                subtotal=Decimal('9999999999.99'),
                unit_price=Decimal('99999.99')
            )
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
    
    @pytest.mark.asyncio
    async def test_special_characters_in_names(self):
        """测试名称中的特殊字符"""
        quote = MockQuoteSheet(
            customer_name='测试<客户>&"特殊"',
            project_name='项目/名称\\test'
        )
        items = [
            MockQuoteItem(
                product_name='产品<名称>&"测试"'
            )
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
    
    @pytest.mark.asyncio
    async def test_many_items(self):
        """测试大量商品（性能测试）"""
        quote = MockQuoteSheet()
        items = [
            MockQuoteItem(
                product_name=f'产品_{i}',
                quantity=i + 1,
                subtotal=Decimal(str(100 * (i + 1)))
            )
            for i in range(100)
        ]
        
        excel_bytes = await self.exporter.generate_standard_quote(quote, items)
        assert excel_bytes is not None
        
        # 验证文件不会太大
        assert len(excel_bytes) < 1024 * 1024  # < 1MB
