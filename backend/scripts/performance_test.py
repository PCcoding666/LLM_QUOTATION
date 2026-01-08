#!/usr/bin/env python3
"""
服务层性能测试脚本
测试计费引擎、报价服务和产品筛选服务的性能

使用方法:
    python scripts/performance_test.py --service pricing
    python scripts/performance_test.py --service quote
    python scripts/performance_test.py --service filter
    python scripts/performance_test.py --all
"""
import argparse
import asyncio
import time
import statistics
from decimal import Decimal
from datetime import datetime, timedelta
from uuid import uuid4
from typing import List, Dict, Any
import sys
import os

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services.pricing_engine import PricingEngine, TieredDiscountRule


class PerformanceMetrics:
    """性能指标收集器"""
    
    def __init__(self, name: str):
        self.name = name
        self.times: List[float] = []
        self.errors: int = 0
    
    def record(self, elapsed: float):
        """记录一次执行时间"""
        self.times.append(elapsed)
    
    def record_error(self):
        """记录一次错误"""
        self.errors += 1
    
    def report(self) -> Dict[str, Any]:
        """生成报告"""
        if not self.times:
            return {"name": self.name, "error": "没有测试数据"}
        
        return {
            "name": self.name,
            "total_runs": len(self.times),
            "errors": self.errors,
            "min_ms": round(min(self.times) * 1000, 3),
            "max_ms": round(max(self.times) * 1000, 3),
            "avg_ms": round(statistics.mean(self.times) * 1000, 3),
            "median_ms": round(statistics.median(self.times) * 1000, 3),
            "p95_ms": round(sorted(self.times)[int(len(self.times) * 0.95)] * 1000, 3) if len(self.times) >= 20 else "N/A",
            "p99_ms": round(sorted(self.times)[int(len(self.times) * 0.99)] * 1000, 3) if len(self.times) >= 100 else "N/A",
            "total_time_s": round(sum(self.times), 3),
            "throughput_per_s": round(len(self.times) / sum(self.times), 2) if sum(self.times) > 0 else 0
        }


def test_pricing_engine(iterations: int = 1000) -> PerformanceMetrics:
    """测试计费引擎性能"""
    metrics = PerformanceMetrics("PricingEngine")
    engine = PricingEngine()
    
    # 添加阶梯折扣规则
    engine.add_rule(TieredDiscountRule([
        {"threshold": 10000, "discount": 0.9},
        {"threshold": 100000, "discount": 0.8},
        {"threshold": 1000000, "discount": 0.7}
    ]))
    
    # 测试场景
    test_cases = [
        # 场景1: 简单LLM计费
        {
            "base_price": Decimal("0.04"),
            "context": {
                "product_type": "llm",
                "input_token_price": 0.04,
                "output_token_price": 0.12,
                "input_tokens": 10000,
                "output_tokens": 5000,
                "thinking_mode_ratio": 0,
                "batch_call_ratio": 0
            }
        },
        # 场景2: 带思考模式的LLM计费
        {
            "base_price": Decimal("0.04"),
            "context": {
                "product_type": "llm",
                "input_token_price": 0.04,
                "output_token_price": 0.12,
                "input_tokens": 50000,
                "output_tokens": 20000,
                "thinking_mode_ratio": 0.5,
                "thinking_mode_multiplier": 1.5,
                "batch_call_ratio": 0.3
            }
        },
        # 场景3: 标准产品计费
        {
            "base_price": Decimal("100"),
            "context": {
                "product_type": "standard",
                "quantity": 10,
                "duration_months": 12
            }
        },
        # 场景4: 大量Token计费
        {
            "base_price": Decimal("0.002"),
            "context": {
                "product_type": "llm",
                "input_token_price": 0.002,
                "output_token_price": 0.008,
                "input_tokens": 1000000,
                "output_tokens": 500000,
                "thinking_mode_ratio": 0,
                "batch_call_ratio": 1.0,
                "quantity": 100000
            }
        }
    ]
    
    print(f"\n🔧 测试 PricingEngine ({iterations} 次迭代)...")
    
    for i in range(iterations):
        test_case = test_cases[i % len(test_cases)]
        try:
            start = time.perf_counter()
            result = engine.calculate(test_case["base_price"], test_case["context"])
            elapsed = time.perf_counter() - start
            metrics.record(elapsed)
            
            # 验证结果
            if "final_price" not in result:
                metrics.record_error()
        except Exception as e:
            metrics.record_error()
            if i < 5:  # 只打印前5个错误
                print(f"  错误 {i}: {e}")
    
    return metrics


def test_excel_export_simulation(iterations: int = 100) -> PerformanceMetrics:
    """模拟Excel导出性能测试"""
    metrics = PerformanceMetrics("ExcelExport(模拟)")
    
    print(f"\n📊 测试 Excel导出 ({iterations} 次迭代)...")
    
    # 模拟不同大小的报价单
    item_counts = [5, 10, 20, 50, 100]
    
    for i in range(iterations):
        item_count = item_counts[i % len(item_counts)]
        
        try:
            start = time.perf_counter()
            
            # 模拟数据生成
            items_data = []
            for j in range(item_count):
                items_data.append({
                    "product_name": f"产品_{j}",
                    "quantity": j + 1,
                    "price": Decimal(str(100 * (j + 1)))
                })
            
            # 模拟计算总价
            total = sum(item["price"] * item["quantity"] for item in items_data)
            
            # 模拟文件生成（实际测试时会导入openpyxl）
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "测试"
                
                # 写入数据
                for idx, item in enumerate(items_data, 1):
                    ws.cell(row=idx, column=1, value=item["product_name"])
                    ws.cell(row=idx, column=2, value=item["quantity"])
                    ws.cell(row=idx, column=3, value=float(item["price"]))
                
                # 保存到内存
                from io import BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                _ = buffer.getvalue()
            except ImportError:
                # 如果没有openpyxl，模拟延迟
                time.sleep(0.001 * item_count)
            
            elapsed = time.perf_counter() - start
            metrics.record(elapsed)
        except Exception as e:
            metrics.record_error()
            if i < 5:
                print(f"  错误 {i}: {e}")
    
    return metrics


def test_filter_simulation(iterations: int = 500) -> PerformanceMetrics:
    """模拟产品筛选性能测试"""
    metrics = PerformanceMetrics("ProductFilter(模拟)")
    
    print(f"\n🔍 测试 产品筛选 ({iterations} 次迭代)...")
    
    # 模拟产品数据
    products = [
        {
            "code": f"prod_{i}",
            "name": f"产品名称_{i}",
            "vendor": ["aliyun", "volcano", "baidu"][i % 3],
            "category": ["AI-大模型-文本生成", "AI-大模型-视觉理解", "AI-大模型-语音"][i % 3],
            "status": "active"
        }
        for i in range(1000)
    ]
    
    # 测试场景
    filter_cases = [
        {"vendor": "aliyun"},
        {"keyword": "产品名称_5"},
        {"vendor": "volcano", "keyword": "产品"},
        {"category": "AI-大模型-文本生成"},
        {}  # 无筛选
    ]
    
    for i in range(iterations):
        filters = filter_cases[i % len(filter_cases)]
        
        try:
            start = time.perf_counter()
            
            # 模拟筛选逻辑
            result = products
            
            if "vendor" in filters:
                result = [p for p in result if p["vendor"] == filters["vendor"]]
            
            if "keyword" in filters:
                kw = filters["keyword"].lower()
                result = [p for p in result if kw in p["name"].lower() or kw in p["code"].lower()]
            
            if "category" in filters:
                result = [p for p in result if p["category"] == filters["category"]]
            
            # 分页
            page_size = 20
            result = result[:page_size]
            
            elapsed = time.perf_counter() - start
            metrics.record(elapsed)
        except Exception as e:
            metrics.record_error()
            if i < 5:
                print(f"  错误 {i}: {e}")
    
    return metrics


def test_quote_calculation(iterations: int = 500) -> PerformanceMetrics:
    """测试报价计算性能"""
    metrics = PerformanceMetrics("QuoteCalculation")
    engine = PricingEngine()
    
    print(f"\n💰 测试 报价计算 ({iterations} 次迭代)...")
    
    # 模拟不同规模的报价单
    item_counts = [3, 5, 10, 20, 50]
    
    for i in range(iterations):
        item_count = item_counts[i % len(item_counts)]
        
        try:
            start = time.perf_counter()
            
            total_original = Decimal("0")
            total_final = Decimal("0")
            
            # 模拟计算每个报价项
            for j in range(item_count):
                context = {
                    "product_type": "llm" if j % 2 == 0 else "standard",
                    "input_token_price": 0.04,
                    "output_token_price": 0.12,
                    "input_tokens": 10000 * (j + 1),
                    "output_tokens": 5000 * (j + 1),
                    "thinking_mode_ratio": 0.2 if j % 3 == 0 else 0,
                    "batch_call_ratio": 0.5 if j % 4 == 0 else 0,
                    "quantity": j + 1,
                    "duration_months": 1
                }
                
                result = engine.calculate(Decimal("0.04"), context)
                total_original += Decimal(str(result["original_price"]))
                total_final += Decimal(str(result["final_price"]))
            
            # 应用全局折扣
            global_discount = Decimal("0.95")
            final_amount = total_final * global_discount
            
            elapsed = time.perf_counter() - start
            metrics.record(elapsed)
        except Exception as e:
            metrics.record_error()
            if i < 5:
                print(f"  错误 {i}: {e}")
    
    return metrics


def print_report(metrics: PerformanceMetrics, threshold_ms: float = 500):
    """打印性能报告"""
    report = metrics.report()
    
    print(f"\n{'='*60}")
    print(f"📈 {report['name']} 性能报告")
    print(f"{'='*60}")
    
    if "error" in report:
        print(f"❌ {report['error']}")
        return
    
    print(f"  总执行次数: {report['total_runs']}")
    print(f"  错误次数: {report['errors']}")
    print(f"  最小耗时: {report['min_ms']} ms")
    print(f"  最大耗时: {report['max_ms']} ms")
    print(f"  平均耗时: {report['avg_ms']} ms")
    print(f"  中位数: {report['median_ms']} ms")
    print(f"  P95: {report['p95_ms']} ms")
    print(f"  P99: {report['p99_ms']} ms")
    print(f"  总耗时: {report['total_time_s']} s")
    print(f"  吞吐量: {report['throughput_per_s']} ops/s")
    
    # 检查是否满足性能要求
    if isinstance(report['avg_ms'], (int, float)) and report['avg_ms'] < threshold_ms:
        print(f"\n  ✅ 性能达标 (平均耗时 < {threshold_ms}ms)")
    else:
        print(f"\n  ⚠️ 性能需关注 (平均耗时 >= {threshold_ms}ms)")


def main():
    parser = argparse.ArgumentParser(description="服务层性能测试")
    parser.add_argument("--service", choices=["pricing", "quote", "filter", "excel"],
                        help="指定要测试的服务")
    parser.add_argument("--all", action="store_true", help="测试所有服务")
    parser.add_argument("--iterations", type=int, default=500, help="迭代次数")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("🚀 报价侠系统 - 服务层性能测试")
    print(f"⏱️  开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    results = []
    
    if args.all or args.service == "pricing":
        metrics = test_pricing_engine(args.iterations)
        print_report(metrics)
        results.append(metrics)
    
    if args.all or args.service == "quote":
        metrics = test_quote_calculation(args.iterations)
        print_report(metrics)
        results.append(metrics)
    
    if args.all or args.service == "filter":
        metrics = test_filter_simulation(args.iterations)
        print_report(metrics)
        results.append(metrics)
    
    if args.all or args.service == "excel":
        metrics = test_excel_export_simulation(min(args.iterations, 100))
        print_report(metrics, threshold_ms=1000)  # Excel导出允许更长时间
        results.append(metrics)
    
    if not args.all and not args.service:
        print("\n请指定 --service 或 --all 参数")
        parser.print_help()
        return
    
    # 汇总报告
    print("\n" + "=" * 60)
    print("📊 测试汇总")
    print("=" * 60)
    
    all_passed = True
    for metrics in results:
        report = metrics.report()
        if "error" not in report:
            status = "✅" if report['avg_ms'] < 500 else "⚠️"
            if report['avg_ms'] >= 500:
                all_passed = False
            print(f"  {status} {report['name']}: {report['avg_ms']}ms (avg), {report['throughput_per_s']} ops/s")
    
    print(f"\n⏱️  结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if all_passed:
        print("\n✅ 所有测试通过性能要求!")
    else:
        print("\n⚠️ 部分测试需要性能优化")


if __name__ == "__main__":
    main()
